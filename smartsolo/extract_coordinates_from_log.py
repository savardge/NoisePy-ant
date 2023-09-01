import os
from openpyxl import Workbook, load_workbook
import statistics

# path to the dataset:
path = "/media/savardg/sandisk4TB/DCCDATA/Riehen/Riehen/"

# preparation of the excel file containing old and new names:
#workbook = load_workbook(filename="position_names.xlsx")
#sheet = workbook.active
#old_names_new_names=[]
#for row in sheet.iter_rows(min_row=2, min_col=1, max_row=204, max_col=2):
#    for cell in row:
#        #print(cell.value, end=" ")
#        old_names_new_names.append(str(cell.value))

fname = "precise_coordinates_mode.txt"

# Create and prepare the resulting txt file:
f = open(fname, "w+") # w for write --> file or create if not exists
f.write(("Name,Longitude,Latitude,Elevation" ) + "\n")
f.close

# Start the loop:
count = 0
# Go into one directory: (ex: 453006384)
old_station_path = os.listdir(path)
for station in old_station_path:
    station_folder_path = path + "/" + str(station)                # F:\\DCCDATA\\Vulcano\\VulcaNODES\\453006384
    stafolder = os.listdir(station_folder_path)
    subfolder_path = station_folder_path + "/" + str(stafolder[0]) # F:\\DCCDATA\\Vulcano\\VulcaNODES\\453006384\\20211114102336
    subfolder = os.listdir(subfolder_path)
    file_path = subfolder_path + "/DigiSolo.LOG"                   # F:\\DCCDATA\\Vulcano\\VulcaNODES\\453006384\\20211114102336\\DigiSolo.LOG
    # Prepare the lists:
    latitudes = []
    longitudes = []
    elevations = []
    # Open the file with the infos:
    with open(file_path) as f:
        for line in f:
            # Get the latitude
            if line.startswith("Latitude"):
                lat = float(line[13:24])
                latitudes.append(lat)
            # Get the longitude
            elif line.startswith("Longitude"):
                lon = float(line[12:24])
                longitudes.append(lon)
                # Get the elevation
            elif line.startswith("Altitude"):
                try: # because some altitudes are "unknown"
                    el = float(line[13:24])
                except:
                    continue
                elevations.append(el)
            else:
                continue
    # Get the mean values for each list
#    latitude = statistics.mean(latitudes)
#    longitude = statistics.mean(longitudes)
#    elevation = statistics.mean(elevations)
    latitude = statistics.mode(latitudes)
    longitude = statistics.mode(longitudes)
    elevation = statistics.mode(elevations)
    # Get the station new name:
    #for old_name in old_names_new_names:
    #    if old_name == station:
    #        index = old_names_new_names.index(old_name)
    #        new_name = old_names_new_names[index+1]
    #Check:
    print("Folder: ", str(station))
    #print("New name: ", str(new_name))
    print("Longitude: ", str(longitude)[:11])
    print("Latitude: ", str(latitude)[:11])
    print("Elevation: ", str(elevation)[:7])
    # Put these info into the created text file:
    f = open(fname, "a+") # a for append
#    f.write((str(new_name) + "," + str(longitude)[:10] + "," + str(latitude)[:11] + "," + str(elevation)[:7] ) + "\n")
    f.write((str(station) + "," + str(longitude)[:10] + "," + str(latitude)[:11] + "," + str(elevation)[:7] ) + "\n")
    f.close
    count = count + 1
    print("-----------------------------------------------------")

print(count, "folders done.")
